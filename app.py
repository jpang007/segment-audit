#!/usr/bin/env python3
"""
Gateway API-Only Audit Dashboard
Testing Gateway API as replacement for Public API
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from flask_cors import CORS
import os
import json
import requests
from pathlib import Path
from datetime import datetime, timedelta
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.secret_key = os.urandom(24)
CORS(app)

# Session configuration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True

# Feature flags
ENABLE_EXPERIMENTAL_FEATURES = os.getenv('ENABLE_EXPERIMENTAL_FEATURES', 'false').lower() == 'true'

# Configuration
DATA_DIR = Path('./audit_data')
DATA_DIR.mkdir(exist_ok=True)

# Global status for audit progress
audit_status = {
    'running': False,
    'progress': 0,
    'message': 'Ready',
    'complete': False,
    'error': None
}

class GatewayAPIClient:
    """Gateway API GraphQL client"""

    def __init__(self, gateway_token, workspace_slug):
        self.gateway_token = gateway_token
        self.workspace_slug = workspace_slug
        self.base_url = "https://app.segment.com/gateway-api/graphql"
        self.headers = {
            "Authorization": f"Bearer {gateway_token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "x-requested-with": "fetch"
        }

    def _execute_query(self, query, variables, max_retries=3):
        """Execute a GraphQL query with retry logic for rate limiting"""
        import time

        for attempt in range(max_retries):
            try:
                payload = {"query": query, "variables": variables}
                response = requests.post(self.base_url, headers=self.headers, json=payload, timeout=30)

                data = response.json()

                if response.status_code == 401:
                    raise Exception("Gateway API authentication failed. Token may be expired.")
                elif response.status_code == 429:
                    # Rate limit - retry with exponential backoff
                    if attempt < max_retries - 1:
                        wait_time = (2 ** attempt) * 5  # 5s, 10s, 20s
                        print(f"\n⚠️  Rate limited. Waiting {wait_time}s before retry {attempt + 1}/{max_retries}...", flush=True)
                        time.sleep(wait_time)
                        continue
                    else:
                        print(f"\n❌ Rate limit exceeded after {max_retries} attempts", flush=True)
                        raise Exception(f"Gateway API rate limit exceeded: {data}")
                elif response.status_code != 200:
                    # Try to get error details from response body
                    print(f"\n=== HTTP {response.status_code} ERROR ===", flush=True)
                    print(f"Response: {json.dumps(data, indent=2)}", flush=True)
                    raise Exception(f"Gateway API returned {response.status_code}: {data}")

                if 'errors' in data:
                    errors = data['errors']
                    error_msg = errors[0].get('message', str(errors)) if errors else 'Unknown error'
                    print(f"\n=== GraphQL ERROR ===", flush=True)
                    print(f"Message: {error_msg}", flush=True)
                    print(f"Full errors: {json.dumps(errors, indent=2)}", flush=True)
                    raise Exception(f"GraphQL error: {error_msg}")

                return data.get('data', {})

            except requests.exceptions.RequestException as e:
                raise Exception(f"Request failed: {str(e)}")

    def get_workspace_connections(self):
        """Get all sources and destinations with connections"""
        query = """
        query getWorkspaceConnections($workspaceSlug: Slug!) {
          workspace(slug: $workspaceSlug) {
            id
            slug
            sources {
              id
              name
              slug
              status
              url
              connectedDestinations {
                id
                __typename
              }
              writeKeys
              flags {
                javascript
                __typename
              }
              settings
              metadata {
                id
                category
                name
                slug
                logos {
                  mark
                  default
                  __typename
                }
                type
                isCloudSource
                options
                __typename
              }
              __typename
            }
            destinations {
              id
              ... on Integration {
                name
                enabled
                __typename
                integrationStatus: status
                connectedSources {
                  id
                  __typename
                }
                metadata {
                  id
                  name
                  type
                  slug
                  logos {
                    mark
                    default
                    __typename
                  }
                  platforms {
                    server
                    browser
                    mobile
                    warehouse
                    __typename
                  }
                  integrationCategories: categories
                  __typename
                }
              }
              ... on Warehouse {
                name
                enabled
                __typename
                warehouseStatus: status
                connectedSources {
                  id
                  __typename
                }
                metadata {
                  id
                  name
                  slug
                  logos {
                    mark
                    default
                    __typename
                  }
                  __typename
                }
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {"workspaceSlug": self.workspace_slug}
        data = self._execute_query(query, variables)
        return data.get('workspace', {})

    def get_all_sources(self):
        """Get all sources with destinations via AllSources query"""
        query = """
        query AllSources($workspaceSlug: Slug!, $cursor: SourcesCursorInput!, $fetchStatus: Boolean!, $fetchDestinations: Boolean!) {
          workspace(slug: $workspaceSlug) {
            id
            sources: sourcesV2(cursor: $cursor) {
              data {
                id
                name
                slug
                status @include(if: $fetchStatus)
                createdAt
                flags {
                  javascript
                  autoInstrumentation
                  __typename
                }
                labels: labelsV2 {
                  key
                  value
                  __typename
                }
                metadata {
                  id
                  name
                  slug
                  logos {
                    mark
                    default
                    __typename
                  }
                  categories
                  category
                  isCloudEventSource
                  __typename
                }
                reverseEtlModels {
                  id
                  enabled
                  __typename
                }
                integrations @include(if: $fetchDestinations) {
                  id
                  name
                  metadata {
                    id
                    name
                    logos {
                      mark
                      default
                      __typename
                    }
                    __typename
                  }
                  __typename
                }
                warehouses @include(if: $fetchDestinations) {
                  id
                  name
                  metadata {
                    id
                    name
                    logos {
                      mark
                      default
                      __typename
                    }
                    __typename
                  }
                  __typename
                }
                __typename
              }
              cursor {
                next
                previous
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug,
            "cursor": {"limit": 200},
            "fetchStatus": True,
            "fetchDestinations": True
        }

        data = self._execute_query(query, variables)
        workspace = data.get('workspace', {})
        sources_result = workspace.get('sources', {})
        return sources_result.get('data', [])

    def get_audiences_with_folders(self, space_id):
        """Get audiences with destinations via AudiencesAndFolders query with folder recursion and pagination"""
        query = """
        query AudiencesAndFolders($workspaceSlug: Slug!, $spaceId: String!, $cursor: RecordCursorInput!, $folderId: String) {
          workspace(slug: $workspaceSlug) {
            id
            space(id: $spaceId) {
              id
              audiencesAndFolders(cursor: $cursor, folderId: $folderId) {
                cursor {
                  limit
                  next
                  hasMore
                  __typename
                }
                data {
                  __typename
                  ... on Folder {
                    folderId: id
                    displayName
                    audienceCount
                  }
                  ... on Audience {
                    audienceId: id
                    key
                    name
                    enabled
                    collection
                    size
                    status
                    destinations {
                      id
                      name
                      enabled
                    }
                  }
                }
              }
            }
          }
        }
        """

        audience_map = {}
        folders_to_process = []

        # Step 1: Get top-level items (with pagination)
        cursor_token = None
        has_more = True
        page_num = 1

        while has_more:
            variables = {
                "workspaceSlug": self.workspace_slug,
                "spaceId": space_id,
                "cursor": {"limit": 200}
            }
            if cursor_token:
                variables["cursor"]["next"] = cursor_token

            data = self._execute_query(query, variables)
            workspace = data.get('workspace', {})
            space = workspace.get('space', {})
            audiences_and_folders = space.get('audiencesAndFolders', {})
            items = audiences_and_folders.get('data', [])
            cursor_info = audiences_and_folders.get('cursor', {})

            print(f"    Page {page_num}: Found {len(items)} items at top level")

            for item in items:
                typename = item.get('__typename', '')

                if typename == 'Folder':
                    folder_id = item.get('folderId')
                    folder_name = item.get('displayName', '')
                    folder_count = item.get('audienceCount', 0)
                    if folder_id and folder_count > 0:
                        folders_to_process.append({
                            'id': folder_id,
                            'name': folder_name,
                            'count': folder_count
                        })
                elif typename in ['Audience', 'RealtimeAudience']:
                    audience_id = item.get('audienceId')
                    if audience_id:
                        destinations = item.get('destinations', []) or []
                        audience_map[audience_id] = {
                            'id': audience_id,
                            'name': item.get('name', ''),
                            'key': item.get('key', ''),
                            'enabled': item.get('enabled', False),
                            'size': item.get('size', 0),
                            'collection': item.get('collection', ''),
                            'status': item.get('status', ''),
                            'destinations': [d.get('name', '') for d in destinations if d.get('name')],
                            'destination_count': len(destinations)
                        }

            # Check for more pages
            has_more = cursor_info.get('hasMore', False)
            cursor_token = cursor_info.get('next') if has_more else None
            page_num += 1

        # Step 2: Query each folder (with pagination)
        for folder in folders_to_process:
            print(f"    Processing folder: {folder['name']} ({folder['count']} audiences)")
            folder_cursor = None
            folder_has_more = True
            folder_page = 1

            while folder_has_more:
                variables = {
                    "workspaceSlug": self.workspace_slug,
                    "spaceId": space_id,
                    "folderId": folder['id'],
                    "cursor": {"limit": 200}
                }
                if folder_cursor:
                    variables["cursor"]["next"] = folder_cursor

                try:
                    data = self._execute_query(query, variables)
                    workspace = data.get('workspace', {})
                    space = workspace.get('space', {})
                    audiences_and_folders = space.get('audiencesAndFolders', {})
                    items = audiences_and_folders.get('data', [])
                    cursor_info = audiences_and_folders.get('cursor', {})

                    print(f"      Folder page {folder_page}: {len(items)} items")

                    for item in items:
                        if item.get('__typename') in ['Audience', 'RealtimeAudience']:
                            audience_id = item.get('audienceId')
                            if audience_id:
                                destinations = item.get('destinations', []) or []
                                audience_map[audience_id] = {
                                    'id': audience_id,
                                    'name': item.get('name', ''),
                                    'key': item.get('key', ''),
                                    'enabled': item.get('enabled', False),
                                    'size': item.get('size', 0),
                                    'collection': item.get('collection', ''),
                                    'status': item.get('status', ''),
                                    'destinations': [d.get('name', '') for d in destinations if d.get('name')],
                                    'destination_count': len(destinations),
                                    'folder': folder['name']
                                }

                    # Check for more pages in folder
                    folder_has_more = cursor_info.get('hasMore', False)
                    folder_cursor = cursor_info.get('next') if folder_has_more else None
                    folder_page += 1

                except Exception as e:
                    print(f"      Error fetching folder {folder['name']} page {folder_page}: {e}")
                    break

        print(f"    Total unique audiences collected: {len(audience_map)}")
        return list(audience_map.values())

    def get_spaces(self):
        """Get all spaces in workspace"""
        query = """
        query GetSpaces($workspaceSlug: Slug!) {
          workspace(slug: $workspaceSlug) {
            id
            spaces {
              id
              slug
              name
              __typename
            }
            __typename
          }
        }
        """

        variables = {"workspaceSlug": self.workspace_slug}
        data = self._execute_query(query, variables)
        workspace = data.get('workspace', {})
        return workspace.get('spaces', [])

    def get_source_schema(self, source_slug):
        """Get source event schema with object properties (traits)"""
        query = """
        query getSourceSchemaWithObjectProperties($workspaceSlug: Slug!, $sourceSlug: String!, $timeframe: Int!, $useFlatClickHouseCounts: Boolean!) {
          workspace(slug: $workspaceSlug) {
            id
            source(slug: $sourceSlug) {
              id
              name
              events(days: $timeframe) @include(if: $useFlatClickHouseCounts) {
                name
                type
                counts @include(if: $useFlatClickHouseCounts) {
                  allowed
                  denied
                  __typename
                }
                __typename
              }
              schema {
                id
                collections {
                  id
                  name
                  events {
                    id
                    name
                    enabled
                    archived
                    createdAt
                    isPlanned
                    __typename
                  }
                  objectProperties {
                    id
                    key
                    type
                    enabled
                    archived
                    lastSeenAt
                    createdAt
                    isPlanned
                    stats(days: $timeframe) {
                      allowed
                      denied
                      __typename
                    }
                    __typename
                  }
                  __typename
                }
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug,
            "sourceSlug": source_slug,
            "timeframe": 7,
            "useFlatClickHouseCounts": True
        }

        data = self._execute_query(query, variables)
        source = data.get('workspace', {}).get('source', {})

        return {
            'events': source.get('events', []),
            'collections': source.get('schema', {}).get('collections', [])
        }

    def get_audience_definition(self, space_id, audience_id):
        """Get audience definition"""
        query = """
        query GetAudienceDefinition($workspaceSlug: Slug!, $spaceId: String!, $audienceId: String!) {
          workspace(slug: $workspaceSlug) {
            space(id: $spaceId) {
              audience(id: $audienceId) {
                id
                name
                key
                engine
                status
                size
                createdAt
                definition {
                  type
                  options
                  __typename
                }
                __typename
              }
            }
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug,
            "spaceId": space_id,
            "audienceId": audience_id
        }

        data = self._execute_query(query, variables)
        audience = data.get('workspace', {}).get('space', {}).get('audience', {})

        return {
            'audience': audience,
            'definition': audience.get('definition', {})
        }

    def get_journeys(self, space_id):
        """Get all journeys and campaigns for a space"""
        query = """
        query GetJourneys($workspaceSlug: Slug!, $spaceId: String!, $cursor: RecordCursorInput!) {
          workspace(slug: $workspaceSlug) {
            space(id: $spaceId) {
              campaignSearch(spaceId: $spaceId, cursor: $cursor) {
                cursor {
                  hasMore
                  next
                }
                data {
                  __typename
                  ... on Campaign {
                    containerId
                    version
                    versionCount
                    updatedAt
                    name
                    variant
                    state
                    definition
                    createdBy {
                      id
                      name
                    }
                    hasPublishedVersion
                    campaignsDestinations: destinations {
                      id
                      metadataId
                      name
                      logo
                    }
                  }
                  ... on Journey {
                    id
                    name
                    description
                    containerId
                    version
                    maxVersion
                    spaceId
                    executionState
                    status
                    emailActionCount
                    smsActionCount
                    whatsAppActionCount
                    mobilePushActionCount
                    destinations {
                      id
                      metadata {
                        ... on IntegrationMetadata {
                          id
                          name
                        }
                        ... on WarehouseMetadata {
                          id
                          name
                        }
                      }
                    }
                    createdBy {
                      id
                      name
                    }
                    updatedAt
                    container {
                      maxVersion
                      isLocked
                      updatedAt
                    }
                  }
                }
              }
            }
          }
        }
        """

        all_items = []
        next_cursor = None

        while True:
            variables = {
                "workspaceSlug": self.workspace_slug,
                "spaceId": space_id,
                "cursor": {
                    "limit": 200,
                    "next": next_cursor
                } if next_cursor else {
                    "limit": 200
                }
            }

            data = self._execute_query(query, variables)
            campaign_search = data.get('workspace', {}).get('space', {}).get('campaignSearch', {})

            if not campaign_search:
                break

            data_items = campaign_search.get('data', [])
            all_items.extend(data_items)

            cursor_info = campaign_search.get('cursor', {})
            has_more = cursor_info.get('hasMore', False)
            next_cursor = cursor_info.get('next')

            if not has_more or not next_cursor:
                break

        return all_items

    def get_identity_resolution_config(self, space_id):
        """Get identity resolution configuration for a space"""
        query = """
        query IdentityResolutionConfig($workspaceSlug: Slug!, $spaceId: String!) {
          workspace(slug: $workspaceSlug) {
            id
            space(id: $spaceId) {
              id
              identityResolutionConfig {
                idTypePriority
                externalIdConfigs {
                  id
                  idType
                  limit
                  mergedLimit
                  mergedLimitTimeRange
                  enforceUnique
                  seen
                  __typename
                }
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug,
            "spaceId": space_id
        }

        try:
            data = self._execute_query(query, variables)
            space_data = data.get('workspace', {}).get('space', {})

            if not space_data:
                print(f"    -> No space data returned for {space_id}")
                return []

            config = space_data.get('identityResolutionConfig', {})

            if not config:
                print(f"    -> No identityResolutionConfig found")
                return []

            # Get the priority order array
            id_type_priority = config.get('idTypePriority', [])
            configs = config.get('externalIdConfigs', [])

            # Add priority to each config based on its position in idTypePriority array
            for conf in configs:
                id_type = conf.get('idType', '')
                if id_type in id_type_priority:
                    conf['priority'] = id_type_priority.index(id_type) + 1  # 1-based priority
                else:
                    conf['priority'] = None

            # Sort by priority
            configs.sort(key=lambda x: x.get('priority') or 999)

            print(f"    -> Found {len(configs)} identity configs")
            if configs:
                print(f"    -> Full config data:")
                for conf in configs:
                    print(f"       {conf.get('idType')}: priority={conf.get('priority')}, limit={conf.get('limit')}, mergedLimit={conf.get('mergedLimit')}, mergedLimitTimeRange={conf.get('mergedLimitTimeRange')}")
            return configs
        except Exception as e:
            print(f"    -> Exception in get_identity_resolution_config: {e}")
            import traceback
            traceback.print_exc()
            return []

    def get_personas_data(self):
        """Get workspace-level personas/profiles entitlements"""
        query = """
        query PersonasData($workspaceSlug: Slug!) {
          workspace(slug: $workspaceSlug) {
            id
            entitiesEnabled: gateOpen(family: "releases", name: "entities")
            flags {
              canSeePersonas
              __typename
            }
            spaces {
              id
              slug
              __typename
            }
            billing {
              isFreeAccount
              __typename
            }
            entitlements {
              quotas {
                connections {
                  linkedEventsEntities
                  linkedEventsRows
                  __typename
                }
                __typename
              }
              features {
                personas
                profiles
                linkedAudiences
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug
        }

        data = self._execute_query(query, variables)
        return data.get('workspace', {})

    def get_profiles_actions(self, space_id, action_type='VIOLATION', days_back=7):
        """Get profile actions (violations/errors) for a space"""
        query = """
        query personasProfilesActions($workspaceSlug: Slug!, $spaceId: String!, $filter: ProfilesActionsFilter!) {
          workspace(slug: $workspaceSlug) {
            id
            space(id: $spaceId) {
              id
              profilesActions(filter: $filter) {
                name
                messageId
                time
                type
                sourceId
                messageSentAt
                messageReceivedAt
                incomingEventType
                displayName
                messageSegmentId
                associatedProfileIds
                collection
                externalIds {
                  type
                  values
                  __typename
                }
                externalIdTypeCausingViolation
                droppedLink
                limitExceededTraits {
                  key
                  value
                  __typename
                }
                limitExceededEvents {
                  name
                  sourceId
                  __typename
                }
                droppedIdentifiers {
                  type
                  id
                  __typename
                }
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        # Calculate time range (last N days)
        end_time = datetime.now()
        start_time = end_time - timedelta(days=days_back)

        # Convert to milliseconds timestamp as strings
        end_ms = str(int(end_time.timestamp() * 1000))
        start_ms = str(int(start_time.timestamp() * 1000))

        variables = {
            "workspaceSlug": self.workspace_slug,
            "spaceId": space_id,
            "filter": {
                "action": {
                    "type": action_type
                },
                "startTime": start_ms,
                "endTime": end_ms
            }
        }

        data = self._execute_query(query, variables)
        space_data = data.get('workspace', {}).get('space', {})
        return space_data.get('profilesActions', [])

    def get_space_sources(self, space_id):
        """Get sources feeding into a specific space with destinations"""
        query = """
        query GetSpaceSources($workspaceSlug: Slug!, $spaceId: String!) {
          workspace(slug: $workspaceSlug) {
            id
            space(id: $spaceId) {
              id
              slug
              name
              sources {
                id
                name
                slug
                status
                metadata {
                  id
                  name
                  category
                  __typename
                }
                integrations {
                  id
                  name
                  __typename
                }
                warehouses {
                  id
                  name
                  __typename
                }
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug,
            "spaceId": space_id
        }

        data = self._execute_query(query, variables)
        space_data = data.get('workspace', {}).get('space', {})
        return space_data.get('sources', [])

    def get_all_destinations(self):
        """Get all destinations (integrations and warehouses) with detailed information"""
        query = """
        query AllDestinations($workspaceSlug: Slug!, $cursor: DestinationsCursorInput!, $filter: DestinationsFilter, $search: DestinationsSearch, $sort: DestinationsSort) {
          workspace(slug: $workspaceSlug) {
            id
            destinations: destinationsV2(
              search: $search
              filter: $filter
              sort: $sort
              cursor: $cursor
            ) {
              data {
                id
                __typename
                ... on Integration {
                  name
                  integrationStatus: status
                  createdAt
                  source {
                    id
                    name
                    slug
                    metadata {
                      id
                      name
                      category
                      categories
                      logos {
                        mark
                        default
                        __typename
                      }
                      __typename
                    }
                    __typename
                  }
                  metadata {
                    id
                    name
                    slug
                    logos {
                      mark
                      default
                      __typename
                    }
                    categories
                    __typename
                  }
                  subscriptions {
                    enabled
                    __typename
                  }
                  enabled
                  __typename
                }
                ... on Warehouse {
                  name
                  warehouseStatus: status
                  createdAt
                  projectOverrides
                  metadata {
                    id
                    name
                    slug
                    logos {
                      mark
                      default
                      __typename
                    }
                    __typename
                  }
                  rawSettings
                  enabled
                  __typename
                }
              }
              cursor {
                next
                previous
                __typename
              }
              __typename
            }
            spaces {
              id
              name
              sources {
                id
                slug
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        all_destinations = []
        cursor = ""

        while True:
            variables = {
                "workspaceSlug": self.workspace_slug,
                "cursor": {
                    "next": cursor,
                    "limit": 100
                },
                "filter": {
                    "category": "ALL"
                }
            }

            data = self._execute_query(query, variables)
            workspace = data.get('workspace', {})
            destinations_response = workspace.get('destinations', {})
            destinations_data = destinations_response.get('data', [])

            all_destinations.extend(destinations_data)

            # Check if there's more data
            next_cursor = destinations_response.get('cursor', {}).get('next', '')
            if not next_cursor:
                break
            cursor = next_cursor

        return {
            'destinations': all_destinations,
            'spaces': workspace.get('spaces', [])
        }

    def get_usage_period_data(self):
        """Get workspace billing, usage, and entitlements data"""
        query = """
        query app__getUsagePeriodData($workspaceSlug: Slug!) {
          workspace(slug: $workspaceSlug) {
            id
            billing {
              canUpgrade
              planName
              start: subscriptionStart
              end: subscriptionEnd
              hasThroughput
              isOnApiPlan
              isOnCalendarMonth
              isOnMtuPlan
              isOnWarehousesPlan
              isOnTwilioDeveloperPlan
              isOnInvoicedPlan
              mtuStrategy
              personasPlan
              profilesPlan
              isAutomaticallyChargedForOverages
              personasComputeCredits
              personasComputeHistory
              personasSyncFrequency
              personasProfileApi
              quota
              usage {
                mtus {
                  anonymous
                  users
                  __typename
                }
                __typename
              }
              overages {
                costPerUnit
                value
                low
                high
                __typename
              }
              startupProgramPromotion {
                grantAmount
                balanceRemaining
                createdAt
                expiresAt
                __typename
              }
              __typename
            }
            historicalBilling {
              planName
              start: subscriptionStart
              end: subscriptionEnd
              hasThroughput
              isOnApiPlan
              isOnCalendarMonth
              isOnInvoicedPlan
              isOnMtuPlan
              isOnWarehousesPlan
              isAutomaticallyChargedForOverages
              mtuStrategy
              usage {
                mtus {
                  anonymous
                  users
                  __typename
                }
                __typename
              }
              overages {
                costPerUnit
                value
                low
                high
                __typename
              }
              quota {
                calls
                rows
                throughput
                trackedObjects
                __typename
              }
              __typename
            }
            entitlements {
              features {
                personas
                functionsExecute
                unifyPlus
                twilioEngage
                profiles
                reverseEtl
                linkedAudiences
                __typename
              }
              quotas {
                global {
                  mtu
                  throughput
                  api {
                    inbound
                    outbound
                    __typename
                  }
                  __typename
                }
                personas {
                  computeCredits
                  sqlTraits
                  computeHistory
                  syncFrequency
                  profileApi
                  spaces
                  journeys
                  journeyNodes
                  maxJourneys
                  activationEvents
                  __typename
                }
                warehouses {
                  rows
                  __typename
                }
                functions {
                  hoursUntilOverage
                  hoursUntilLockout
                  __typename
                }
                connections {
                  reverseEtlRows
                  linkedEventsRows
                  linkedEventsEntities
                  __typename
                }
                __typename
              }
              __typename
            }
            isFunctionsTermsAccepted
            inFunctionsBeta
            __typename
          }
        }
        """

        variables = {"workspaceSlug": self.workspace_slug}
        data = self._execute_query(query, variables)
        return data.get('workspace', {})

    def get_audit_trail_events(self, page=1, per_page=100):
        """Get audit trail events for governance and activity analysis"""
        query = """
        query app__getAuditTrailEvents($workspaceSlug: Slug!, $pageArgs: PageArgs, $filter: AuditEventFilter!) {
          workspace(slug: $workspaceSlug) {
            id
            auditEvents(pagination: $pageArgs, filter: $filter) {
              pageInfo {
                page
                perPage
                totalEntries
                hasNextPage
                hasPreviousPage
              }
              nodes {
                id
                timestamp
                type
                target
                resource {
                  id
                  type
                  source {
                    id
                    name
                    slug
                  }
                  integration {
                    id
                    name
                    metadata {
                      id
                      slug
                    }
                  }
                  trackingPlan {
                    id
                    name
                    resourceId
                  }
                  destinationFilter {
                    id
                    name
                    sourceSlug
                    destinationMetadataSlug
                  }
                }
                subject {
                  id
                  type
                  user {
                    id
                    name
                    email
                  }
                  token {
                    id
                    description
                  }
                }
              }
            }
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug,
            "pageArgs": {
                "page": page,
                "perPage": per_page
            },
            "filter": {}
        }

        data = self._execute_query(query, variables)
        return data.get('workspace', {}).get('auditEvents', {})

    def get_usage_period_data(self):
        """Get comprehensive billing and usage data including contract dates"""
        query = """
        query app__getUsagePeriodData($workspaceSlug: Slug!) {
          workspace(slug: $workspaceSlug) {
            id
            billing {
              canUpgrade
              planName
              start: subscriptionStart
              end: subscriptionEnd
              hasThroughput
              isOnApiPlan
              isOnCalendarMonth
              isOnMtuPlan
              isOnWarehousesPlan
              isOnTwilioDeveloperPlan
              isOnInvoicedPlan
              isOnBusinessTier
              isOnHighVolumePlan
              mtuStrategy
              personasPlan
              profilesPlan
              isAutomaticallyChargedForOverages
              personasComputeCredits
              personasComputeHistory
              personasSyncFrequency
              personasProfileApi
              quota
              usage {
                mtus {
                  anonymous
                  users
                  __typename
                }
                __typename
              }
              overages {
                costPerUnit
                value
                low
                high
                __typename
              }
              startupProgramPromotion {
                grantAmount
                balanceRemaining
                createdAt
                expiresAt
                __typename
              }
              __typename
            }
            historicalBilling {
              planName
              start: subscriptionStart
              end: subscriptionEnd
              hasThroughput
              isOnApiPlan
              isOnCalendarMonth
              isOnInvoicedPlan
              isOnMtuPlan
              isOnWarehousesPlan
              isAutomaticallyChargedForOverages
              mtuStrategy
              usage {
                mtus {
                  anonymous
                  users
                  __typename
                }
                __typename
              }
              overages {
                costPerUnit
                value
                low
                high
                __typename
              }
              quota {
                calls
                rows
                throughput
                trackedObjects
                __typename
              }
              __typename
            }
            entitlements {
              features {
                personas
                functionsExecute
                unifyPlus
                twilioEngage
                profiles
                reverseEtl
                linkedAudiences
                __typename
              }
              quotas {
                global {
                  mtu
                  throughput
                  api {
                    inbound
                    outbound
                    __typename
                  }
                  __typename
                }
                personas {
                  computeCredits
                  sqlTraits
                  computeHistory
                  syncFrequency
                  profileApi
                  spaces
                  journeys
                  journeyNodes
                  maxJourneys
                  activationEvents
                  __typename
                }
                warehouses {
                  rows
                  __typename
                }
                functions {
                  hoursUntilOverage
                  hoursUntilLockout
                  __typename
                }
                connections {
                  reverseEtlRows
                  linkedEventsRows
                  linkedEventsEntities
                  __typename
                }
                __typename
              }
              __typename
            }
            isFunctionsTermsAccepted
            inFunctionsBeta
            __typename
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug
        }

        data = self._execute_query(query, variables)
        return data.get('workspace', {})

    def get_mtu_data(self, workspace_id, start_date, end_date, is_monthly=True):
        """Get MTU (Monthly Tracked Users) data for billing analysis"""
        query = """
        query app__GetMtuForPeriod($workspaceId: ID!, $isMonthly: Boolean!, $period: Period!, $isOnCalendarMonth: Boolean!, $mtuStrategy: String!) {
          workspace(id: $workspaceId) {
            id
            billing {
              quota
              __typename
            }
            insights {
              mtuPerMonth @include(if: $isMonthly) {
                users
                anonymous
                period
                __typename
              }
              mtuCumulativePerDay(
                period: $period
                isOnCalendarMonth: $isOnCalendarMonth
                mtuStrategy: $mtuStrategy
              ) @skip(if: $isMonthly) {
                users
                anonymous
                period
                __typename
              }
              mtuDailyCumulativePerSource(
                period: $period
                isOnCalendarMonth: $isOnCalendarMonth
                mtuStrategy: $mtuStrategy
              ) @skip(if: $isMonthly) {
                source {
                  id
                  name
                  __typename
                }
                users
                anonymous
                period
                __typename
              }
              sourceUserCountsPerMonth @include(if: $isMonthly) {
                period
                sources {
                  source {
                    id
                    name
                    __typename
                  }
                  users
                  anonymous
                  __typename
                }
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {
            "workspaceId": workspace_id,
            "isMonthly": is_monthly,
            "period": {
                "start": start_date,
                "end": end_date
            },
            "isOnCalendarMonth": True,
            "mtuStrategy": "de-duped"
        }

        data = self._execute_query(query, variables)
        return data.get('workspace', {})

    def get_data_flows(self, pagination_count=10):
        """Get workspace onboarding use cases and data flows"""
        query = """
        fragment DataFlowOutput on DataFlow {
          id
          resources {
            id
            resourceType
            resourceId
            autoAdded
            __typename
          }
          actions {
            id
            action
            __typename
          }
          __typename
        }

        query getWorkspaceOnboardingUseCaseDataFlows($workspaceSlug: Slug!, $pagination: PaginationInput) {
          workspace(slug: $workspaceSlug) {
            id
            spaces {
              id
              slug
              __typename
            }
            entitlements {
              features {
                protocols
                profiles
                personas
                __typename
              }
              __typename
            }
            onboardingUseCases(pagination: $pagination) {
              data {
                id
                businessGoal
                businessGoalDesiredOutcome
                businessGoalIcon
                useCase
                useCaseObjective
                useCaseIcon
                definition {
                  sourceCategories
                  destinationCategories
                  events {
                    name
                    description
                    properties
                    __typename
                  }
                  __typename
                }
                devDataFlow {
                  ...DataFlowOutput
                  __typename
                }
                prodDataFlow {
                  ...DataFlowOutput
                  __typename
                }
                __typename
              }
              pagination {
                current
                next
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {
            "workspaceSlug": self.workspace_slug,
            "pagination": {
                "count": pagination_count
            }
        }

        data = self._execute_query(query, variables)
        return data.get('workspace', {})

    def get_destination_delivery_metrics(self, destination_id, start_time, end_time, workspace_id):
        """Get delivery metrics for a specific destination (success, failures, retries)"""
        query = """
        query getSuccessfullyDelivered($integrationID: String!, $workspaceID: String!, $destinationConfigId: String!, $groupBy: [String!]!, $startTime: Date!, $endTime: Date!, $granularity: Granularity!, $pagination: PaginationInput, $subscriptionId: String, $filterBy: String) {
          integration(id: $integrationID) {
            id
            name
            deliveryOverview {
              successfulDelivery(
                workspaceId: $workspaceID
                startTime: $startTime
                endTime: $endTime
                granularity: $granularity
                destinationConfigId: $destinationConfigId
                groupBy: $groupBy
                pagination: $pagination
                subscriptionId: $subscriptionId
                filterBy: $filterBy
              ) {
                dataset {
                  total
                  totalRetryCount
                  eventType
                  eventName
                  appVersion
                  action
                  series {
                    time
                    count
                    retryCount
                    __typename
                  }
                  __typename
                }
                pagination {
                  current
                  next
                  totalEntries
                  previous
                  __typename
                }
                __typename
              }
              __typename
            }
            __typename
          }
        }
        """

        variables = {
            "integrationID": destination_id,
            "workspaceID": workspace_id,
            "destinationConfigId": destination_id,
            "groupBy": [],
            "startTime": start_time,
            "endTime": end_time,
            "granularity": "day"
        }

        try:
            data = self._execute_query(query, variables)
            return data.get('integration', {})
        except Exception as e:
            print(f"Error fetching delivery metrics for {destination_id}: {e}")
            return None


# ============================================================================
# ROUTES
# ============================================================================

@app.route('/')
def index():
    """Landing page with Gateway token input"""
    return render_template('gateway_index.html')

@app.route('/start-audit', methods=['POST'])
def start_audit():
    """Start audit collection with Gateway API token"""
    try:
        data = request.json
        gateway_token = data.get('gateway_token', '').strip()
        workspace_slug = data.get('workspace_slug', '').strip()
        customer_name = data.get('customer_name', '').strip() or workspace_slug
        fetch_definitions = data.get('fetch_definitions', False)

        if not gateway_token or not workspace_slug:
            return jsonify({'error': 'Gateway token and workspace slug are required'}), 400

        # Store in session
        session['gateway_token'] = gateway_token
        session['workspace_slug'] = workspace_slug
        session['customer_name'] = customer_name
        session.permanent = True

        # Start audit in background
        import threading
        thread = threading.Thread(target=run_audit, args=(gateway_token, workspace_slug, customer_name, fetch_definitions))
        thread.daemon = True
        thread.start()

        return jsonify({'status': 'started'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

def run_audit(gateway_token, workspace_slug, customer_name, fetch_definitions=False):
    """Run the audit collection"""
    global audit_status

    try:
        # Reset all status fields for new audit
        audit_status['running'] = True
        audit_status['progress'] = 0
        audit_status['message'] = 'Initializing Gateway API client...'
        audit_status['error'] = None
        audit_status['complete'] = False

        # Clear old audit data files
        for file_path in DATA_DIR.glob('gateway_*.csv'):
            file_path.unlink()
        for file_path in DATA_DIR.glob('gateway_*.json'):
            file_path.unlink()
        print("Cleared old audit data files")

        client = GatewayAPIClient(gateway_token, workspace_slug)

        # Get spaces
        audit_status['message'] = 'Fetching spaces...'
        audit_status['progress'] = 10
        spaces = client.get_spaces()
        print(f"Found {len(spaces)} spaces")

        # Get sources
        audit_status['message'] = 'Fetching sources...'
        audit_status['progress'] = 30
        sources = client.get_all_sources()
        print(f"Found {len(sources)} sources")

        # Get workspace connections
        audit_status['message'] = 'Fetching source-destination connections...'
        audit_status['progress'] = 50
        connections = client.get_workspace_connections()

        # Get audiences for each space
        audit_status['message'] = 'Collecting audiences...'
        audit_status['progress'] = 70
        all_audiences = []
        print(f"\n=== FETCHING AUDIENCES ===", flush=True)
        print(f"Found {len(spaces)} spaces to query", flush=True)
        for idx, space in enumerate(spaces):
            space_id = space.get('id')
            space_name = space.get('name', 'Unknown')
            print(f"\nQuerying space {idx + 1}/{len(spaces)}: {space_name} (ID: {space_id})", flush=True)
            try:
                audiences = client.get_audiences_with_folders(space_id)
                print(f"  -> Returned {len(audiences)} audiences", flush=True)

                # Optionally fetch definitions for each audience
                for aud_idx, aud in enumerate(audiences):
                    aud['space_id'] = space_id
                    aud['space_name'] = space_name
                    aud_name = aud.get('name', 'Unnamed')
                    print(f"     - {aud_name}", flush=True)

                    if fetch_definitions:
                        aud_id = aud.get('id', '')
                        try:
                            print(f"       Fetching definition for {aud_name}...", flush=True)
                            definition_data = client.get_audience_definition(space_id, aud_id)
                            definition = definition_data.get('definition', {})
                            definition_type = definition.get('type', '')
                            definition_options = definition.get('options', {})

                            # Store definition (SQL or AST)
                            if definition_type == 'segment_sql' and definition_options:
                                aud['definition_sql'] = definition_options.get('sql', '')
                            elif definition_type == 'ast' and definition_options:
                                # Store AST as JSON string for AI export
                                aud['definition_sql'] = json.dumps(definition_options)
                            else:
                                aud['definition_sql'] = ''

                            audit_status['message'] = f'Collecting audiences... ({aud_idx + 1}/{len(audiences)} in {space_name})'

                            # Rate limiting - sleep 2s between API calls to avoid Gateway API limits
                            import time
                            time.sleep(2)
                        except Exception as def_error:
                            print(f"       Failed to fetch definition: {def_error}", flush=True)
                            aud['definition_sql'] = ''

                all_audiences.extend(audiences)
                print(f"  -> Total audiences so far: {len(all_audiences)}", flush=True)
            except Exception as e:
                print(f"  -> ERROR: {e}", flush=True)
                import traceback
                traceback.print_exc()

        # Get journeys for each space
        audit_status['message'] = 'Collecting journeys...'
        audit_status['progress'] = 80
        all_journeys = []
        print(f"\n=== FETCHING JOURNEYS ===")
        for idx, space in enumerate(spaces):
            space_id = space.get('id')
            space_name = space.get('name', 'Unknown')
            print(f"\nQuerying space {idx + 1}/{len(spaces)}: {space_name}")
            try:
                journeys = client.get_journeys(space_id)
                print(f"  -> Returned {len(journeys)} items (journeys + campaigns)")
                for journey in journeys:
                    journey['space_id'] = space_id
                    journey['space_name'] = space_name
                all_journeys.extend(journeys)
                print(f"  -> Total items so far: {len(all_journeys)}")
            except Exception as e:
                print(f"  -> ERROR: {e}")
                print(f"     (This may be normal if workspace doesn't have Engage feature)")

        # Get destinations with usage and data flow context
        audit_status['message'] = 'Collecting destinations...'
        audit_status['progress'] = 82
        print(f"\n=== FETCHING DESTINATIONS ===")
        destinations_data = {}
        usage_data = {}
        data_flows = {}
        try:
            destinations_data = client.get_all_destinations()
            print(f"  -> Found {len(destinations_data.get('destinations', []))} destinations")
        except Exception as e:
            print(f"  -> ERROR fetching destinations: {e}")

        try:
            usage_data = client.get_usage_period_data()
            billing = usage_data.get('billing', {})
            print(f"  -> Billing plan: {billing.get('planName', 'Unknown')}")
            mtus = billing.get('usage', {}).get('mtus', {})
            print(f"  -> MTUs: {mtus.get('users', 0)} users + {mtus.get('anonymous', 0)} anonymous")
        except Exception as e:
            print(f"  -> ERROR fetching usage data: {e}")

        try:
            data_flows = client.get_data_flows()
            use_cases = data_flows.get('onboardingUseCases', {}).get('data', [])
            print(f"  -> Found {len(use_cases)} use cases")
        except Exception as e:
            print(f"  -> ERROR fetching data flows: {e}")

        # Get audit trail events for governance insights (last 90 days)
        audit_status['message'] = 'Collecting audit trail...'
        audit_status['progress'] = 85
        print(f"\n=== FETCHING AUDIT TRAIL ===")
        audit_events = []
        try:
            # Fetch events until we hit 90 days or run out
            # Estimate: 50 pages (5000 events) should cover most workspaces for 90 days
            max_pages = 50
            for page in range(1, max_pages + 1):
                if page % 10 == 0:
                    print(f"  Page {page}/{max_pages}... ({len(audit_events)} events so far)")

                result = client.get_audit_trail_events(page=page, per_page=100)
                nodes = result.get('nodes', [])

                if not nodes:
                    break

                audit_events.extend(nodes)

                # Check if we've reached 90 days back
                if nodes:
                    oldest_event = nodes[-1]
                    event_time = oldest_event.get('timestamp')
                    if event_time:
                        from datetime import datetime, timedelta
                        event_date = datetime.fromisoformat(event_time.replace('Z', '+00:00'))
                        ninety_days_ago = datetime.now(event_date.tzinfo) - timedelta(days=90)

                        if event_date < ninety_days_ago:
                            print(f"  -> Reached 90 days back at page {page}")
                            break

                # Stop if no more pages
                if not result.get('pageInfo', {}).get('hasNextPage', False):
                    print(f"  -> No more pages available")
                    break

            print(f"  -> Collected {len(audit_events)} audit events")
        except Exception as e:
            print(f"  -> ERROR fetching audit trail: {e}")
            print(f"     (Audit trail may not be available on this plan)")

        # Get MTU (Monthly Tracked Users) data
        audit_status['message'] = 'Collecting MTU data...'
        audit_status['progress'] = 87
        print(f"\n=== FETCHING MTU DATA ===")
        mtu_data = {}
        usage_period_data = {}
        try:
            # Get comprehensive usage period data (includes contract dates, billing tier, etc.)
            usage_period_data = client.get_usage_period_data()

            billing = usage_period_data.get('billing', {})
            plan_name = billing.get('planName')
            contract_start = billing.get('start')
            contract_end = billing.get('end')
            is_business = billing.get('isOnBusinessTier')
            is_high_volume = billing.get('isOnHighVolumePlan')

            print(f"  -> Plan: {plan_name}")
            print(f"  -> Contract: {contract_start} to {contract_end}")
            print(f"  -> Business Tier: {is_business}, High Volume: {is_high_volume}")

            # Get workspace ID from workspace_connections
            workspace_data = client.get_workspace_connections()
            workspace_id = workspace_data.get('id')

            if workspace_id:
                # Get last 12 months of MTU data
                from datetime import datetime, timedelta
                end_date = datetime.utcnow()
                start_date = end_date - timedelta(days=365)

                mtu_data = client.get_mtu_data(
                    workspace_id,
                    start_date.isoformat() + 'Z',
                    end_date.isoformat() + 'Z',
                    is_monthly=True
                )

                # Merge usage period data into mtu_data for frontend
                if 'billing' not in mtu_data:
                    mtu_data['billing'] = {}

                # Merge all billing info
                mtu_data['billing'].update(billing)
                mtu_data['entitlements'] = usage_period_data.get('entitlements', {})
                mtu_data['historicalBilling'] = usage_period_data.get('historicalBilling', [])

                quota = mtu_data.get('billing', {}).get('quota')
                mtu_months = mtu_data.get('insights', {}).get('mtuPerMonth', [])
                print(f"  -> MTU Quota: {quota if quota else 'Not set'}")
                print(f"  -> Collected {len(mtu_months)} months of MTU data")
            else:
                print(f"  -> Could not determine workspace ID")
        except Exception as e:
            print(f"  -> ERROR fetching MTU data: {e}")
            print(f"     (MTU data may not be available on this plan)")

        # Get profile insights data
        audit_status['message'] = 'Collecting profile insights...'
        audit_status['progress'] = 90
        print(f"\n=== FETCHING PROFILE INSIGHTS ===")

        # Get workspace-level personas data
        personas_data = {}
        try:
            personas_data = client.get_personas_data()
            print(f"  -> Workspace entitlements: personas={personas_data.get('entitlements', {}).get('features', {}).get('personas')}, profiles={personas_data.get('entitlements', {}).get('features', {}).get('profiles')}")
        except Exception as e:
            print(f"  -> ERROR fetching personas data: {e}")

        # Get identity resolution config and space sources for each space
        all_identity_configs = []
        all_space_sources = []
        for idx, space in enumerate(spaces):
            space_id = space.get('id')
            space_name = space.get('name', 'Unknown')
            print(f"\nQuerying identity config for space {idx + 1}/{len(spaces)}: {space_name}")
            try:
                configs = client.get_identity_resolution_config(space_id)
                if configs:
                    print(f"  -> Returned {len(configs)} external ID types")
                    for config in configs:
                        # Format the limit display
                        merged_limit = config.get('mergedLimit')
                        merged_time_range = config.get('mergedLimitTimeRange')
                        limit_display = config.get('limit', '')
                        seen = config.get('seen', False)

                        # If mergedLimit exists, show it with time range
                        if merged_limit and merged_time_range:
                            # Convert seconds to human readable format
                            seconds = int(merged_time_range)
                            if seconds == 86400:  # 1 day
                                time_period = "daily"
                            elif seconds == 604800:  # 7 days
                                time_period = "weekly"
                            elif seconds == 2592000:  # 30 days
                                time_period = "monthly"
                            elif seconds == 31536000:  # 365 days
                                time_period = "annually"
                            else:
                                # Calculate days if not a standard period
                                days = seconds // 86400
                                time_period = f"{days} days"

                            limit_display = f"{merged_limit} {time_period}"
                        elif merged_limit:
                            limit_display = f"{merged_limit} ever"

                        all_identity_configs.append({
                            'space_id': space_id,
                            'space_name': space_name,
                            'id_type': config.get('idType', ''),
                            'priority': config.get('priority', ''),
                            'limit': limit_display,
                            'seen': 'Yes' if seen else 'No'
                        })
                else:
                    print(f"  -> No identity configs returned")
            except Exception as e:
                print(f"  -> ERROR fetching identity config: {e}")
                import traceback
                traceback.print_exc()

            # Get sources for this space
            try:
                space_sources = client.get_space_sources(space_id)
                print(f"  -> Returned {len(space_sources)} sources")
                for source in space_sources:
                    # Extract destination names from integrations and warehouses
                    destinations = []
                    for integration in source.get('integrations', []):
                        if integration.get('name'):
                            destinations.append(integration['name'])
                    for warehouse in source.get('warehouses', []):
                        if warehouse.get('name'):
                            destinations.append(warehouse['name'])

                    all_space_sources.append({
                        'space_id': space_id,
                        'space_name': space_name,
                        'source_id': source.get('id', ''),
                        'source_name': source.get('name', ''),
                        'source_status': source.get('status', ''),
                        'source_type': source.get('metadata', {}).get('name', ''),
                        'source_category': source.get('metadata', {}).get('category', ''),
                        'destinations': ', '.join(destinations) if destinations else ''
                    })
            except Exception as e:
                print(f"  -> ERROR fetching space sources: {e}")

        # Save data
        audit_status['message'] = 'Saving audit data...'
        audit_status['progress'] = 90

        # Save sources CSV with traits and event count
        if sources:
            with open(DATA_DIR / 'gateway_sources.csv', 'w', newline='', encoding='utf-8') as f:
                fieldnames = ['Workspace', 'ID', 'Name', 'Slug', 'Status', 'Type', 'Technical Type', 'Category', 'Created At', 'Labels', 'Connected Destinations', 'Destination Types', 'Connected Warehouses', 'Warehouse Types', 'Identify Traits', 'Event Count']
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for idx, source in enumerate(sources):
                    labels = ', '.join([f"{l['key']}={l['value']}" for l in source.get('labels', [])])

                    # Format destinations - CSV gets "Name (Type)", UI gets just types
                    integrations_csv = []  # For CSV: "Name (Type)"
                    integrations_types = []  # For UI: just "Type"
                    for i in source.get('integrations', []):
                        dest_name = i.get('name', '')
                        dest_type = i.get('metadata', {}).get('name', '')
                        if dest_name:
                            if dest_type:
                                integrations_csv.append(f"{dest_name} ({dest_type})")
                                integrations_types.append(dest_type)
                            else:
                                integrations_csv.append(dest_name)
                                integrations_types.append(dest_name)

                    integrations_csv_str = ', '.join(integrations_csv)
                    integrations_types_str = ', '.join(integrations_types)

                    # Format warehouses - CSV gets "Name (Type)", UI gets just types
                    warehouses_csv = []
                    warehouses_types = []
                    for w in source.get('warehouses', []):
                        wh_name = w.get('name', '')
                        wh_type = w.get('metadata', {}).get('name', '')
                        if wh_name:
                            if wh_type:
                                warehouses_csv.append(f"{wh_name} ({wh_type})")
                                warehouses_types.append(wh_type)
                            else:
                                warehouses_csv.append(wh_name)
                                warehouses_types.append(wh_name)

                    warehouses_csv_str = ', '.join(warehouses_csv)
                    warehouses_types_str = ', '.join(warehouses_types)

                    # Fetch traits and event count for this source
                    traits_list = []
                    event_count = 0
                    try:
                        audit_status['message'] = f'Fetching schema for {source.get("name", "source")} ({idx + 1}/{len(sources)})...'
                        schema_data = client.get_source_schema(source.get('slug', ''))

                        # Count total events in schema
                        for collection in schema_data.get('collections', []):
                            events = collection.get('events', [])
                            event_count += len(events)

                            # Look for "users" collection (Identify traits)
                            if collection.get('name', '').lower() == 'users':
                                for prop in collection.get('objectProperties', []):
                                    trait_key = prop.get('key', '')
                                    if trait_key:
                                        # Only include enabled, non-archived traits
                                        if prop.get('enabled', True) and not prop.get('archived', False):
                                            traits_list.append(trait_key)

                        # Add delay to avoid rate limits
                        import time
                        time.sleep(1)
                    except Exception as e:
                        print(f"    -> Failed to fetch schema for {source.get('slug', '')}: {e}")

                    traits_str = ', '.join(traits_list) if traits_list else ''

                    # Use category for Type column (more useful: "Website", "Server", "Warehouse", "Custom")
                    # Keep metadata.name as separate field for technical detail
                    metadata_name = source.get('metadata', {}).get('name', '')
                    metadata_category = source.get('metadata', {}).get('category', '')

                    # Type shows category - capitalize first letter (Website, Server, Warehouse, Custom)
                    display_type = metadata_category.capitalize() if metadata_category else metadata_name

                    writer.writerow({
                        'Workspace': workspace_slug,
                        'ID': source.get('id', ''),
                        'Name': source.get('name', ''),
                        'Slug': source.get('slug', ''),
                        'Status': source.get('status', ''),
                        'Type': display_type,  # Shows category: Website, Server, Warehouse, Custom
                        'Category': metadata_category,  # Same as Type for now
                        'Technical Type': metadata_name,  # Javascript, HTTP API, Redshift, etc.
                        'Created At': source.get('createdAt', ''),
                        'Labels': labels,
                        'Connected Destinations': integrations_csv_str,  # Full "Name (Type)" for CSV
                        'Destination Types': integrations_types_str,  # Just "Type" for UI
                        'Connected Warehouses': warehouses_csv_str,  # Full "Name (Type)" for CSV
                        'Warehouse Types': warehouses_types_str,  # Just "Type" for UI
                        'Identify Traits': traits_str,
                        'Event Count': event_count
                    })

        # Save audiences CSV
        if all_audiences:
            with open(DATA_DIR / 'gateway_audiences.csv', 'w', newline='', encoding='utf-8') as f:
                # Include Definition column if definitions were fetched
                fieldnames = ['Workspace', 'ID', 'Name', 'Key', 'Enabled', 'Size', 'Space', 'Space ID', 'Folder', 'Destinations', 'Destination Count']
                if fetch_definitions:
                    fieldnames.append('Definition')

                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for aud in all_audiences:
                    row = {
                        'Workspace': workspace_slug,
                        'ID': aud.get('id', ''),
                        'Name': aud.get('name', ''),
                        'Key': aud.get('key', ''),
                        'Enabled': aud.get('enabled', False),
                        'Size': aud.get('size', 0),
                        'Space': aud.get('space_name', ''),
                        'Space ID': aud.get('space_id', ''),
                        'Folder': aud.get('folder', ''),
                        'Destinations': ', '.join(aud.get('destinations', [])),
                        'Destination Count': aud.get('destination_count', 0)
                    }
                    if fetch_definitions:
                        row['Definition'] = aud.get('definition_sql', '')
                    writer.writerow(row)

        # Save journeys CSV
        if all_journeys:
            with open(DATA_DIR / 'gateway_journeys.csv', 'w', newline='', encoding='utf-8') as f:
                fieldnames = ['Workspace', 'Type', 'ID', 'Name', 'Description', 'State', 'Status', 'Space', 'Space ID',
                             'Current Version', 'Max Version', 'Destinations', 'Destination Count',
                             'Created By', 'Updated At']
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for item in all_journeys:
                    typename = item.get('__typename', '')

                    if typename == 'Journey':
                        destinations = []
                        for dest in item.get('destinations', []):
                            metadata = dest.get('metadata', {})
                            if metadata and isinstance(metadata, dict):
                                destinations.append(metadata.get('name', ''))

                        dest_list = ', '.join(destinations)

                        writer.writerow({
                            'Workspace': workspace_slug,
                            'Type': 'Journey',
                            'ID': item.get('id', ''),
                            'Name': item.get('name', ''),
                            'Description': item.get('description', ''),
                            'State': item.get('status', ''),
                            'Status': item.get('executionState', ''),
                            'Space': item.get('space_name', ''),
                            'Space ID': item.get('space_id', ''),
                            'Current Version': item.get('version', ''),
                            'Max Version': item.get('maxVersion', ''),
                            'Destinations': dest_list,
                            'Destination Count': len(destinations),
                            'Created By': item.get('createdBy', {}).get('name', ''),
                            'Updated At': item.get('updatedAt', '')
                        })

                    elif typename == 'Campaign':
                        destinations = [d.get('name', '') for d in item.get('campaignsDestinations', [])]
                        dest_list = ', '.join(destinations)

                        state = 'published' if item.get('hasPublishedVersion') else 'draft'

                        writer.writerow({
                            'Workspace': workspace_slug,
                            'Type': 'Campaign',
                            'ID': item.get('containerId', ''),
                            'Name': item.get('name', ''),
                            'Description': '',
                            'State': state,
                            'Status': item.get('state', ''),
                            'Space': item.get('space_name', ''),
                            'Space ID': item.get('space_id', ''),
                            'Current Version': item.get('version', ''),
                            'Max Version': item.get('versionCount', ''),
                            'Destinations': dest_list,
                            'Destination Count': len(destinations),
                            'Created By': item.get('createdBy', {}).get('name', ''),
                            'Updated At': item.get('updatedAt', '')
                        })

        # Save profile insights data (always create file even if empty)
        with open(DATA_DIR / 'gateway_profile_insights.csv', 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['Workspace', 'Space', 'Space ID', 'ID Type', 'Priority', 'Limit', 'Seen']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for config in all_identity_configs:
                writer.writerow({
                    'Workspace': workspace_slug,
                    'Space': config.get('space_name', ''),
                    'Space ID': config.get('space_id', ''),
                    'ID Type': config.get('id_type', ''),
                    'Priority': config.get('priority', ''),
                    'Limit': config.get('limit', ''),
                    'Seen': config.get('seen', 'No')
                })

        print(f"\nSaved {len(all_identity_configs)} identity configs to CSV")

        # Save space sources data (always create file even if empty)
        with open(DATA_DIR / 'gateway_space_sources.csv', 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['Workspace', 'Space', 'Space ID', 'Source Name', 'Status', 'Type', 'Category', 'Destinations']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for source in all_space_sources:
                writer.writerow({
                    'Workspace': workspace_slug,
                    'Space': source.get('space_name', ''),
                    'Space ID': source.get('space_id', ''),
                    'Source Name': source.get('source_name', ''),
                    'Status': source.get('source_status', ''),
                    'Type': source.get('source_type', ''),
                    'Category': source.get('source_category', ''),
                    'Destinations': source.get('destinations', '')
                })

        print(f"Saved {len(all_space_sources)} space sources to CSV")

        # Save personas entitlements
        with open(DATA_DIR / 'gateway_personas_entitlements.json', 'w') as f:
            json.dump(personas_data, f, indent=2)

        # Save destinations data
        with open(DATA_DIR / 'gateway_destinations.json', 'w') as f:
            json.dump(destinations_data, f, indent=2)

        with open(DATA_DIR / 'gateway_usage_data.json', 'w') as f:
            json.dump(usage_data, f, indent=2)

        with open(DATA_DIR / 'gateway_data_flows.json', 'w') as f:
            json.dump(data_flows, f, indent=2)

        print(f"Saved destinations, usage, and data flow data")

        # Save audit trail events
        with open(DATA_DIR / 'gateway_audit_trail.json', 'w') as f:
            json.dump(audit_events, f, indent=2)

        print(f"Saved {len(audit_events)} audit trail events")

        # Save MTU data
        with open(DATA_DIR / 'gateway_mtu.json', 'w') as f:
            json.dump(mtu_data, f, indent=2)

        print(f"Saved MTU data")

        # Save summary
        destinations_list = destinations_data.get('destinations', [])
        summary = {
            'audit_date': datetime.now().isoformat(),
            'customer_name': customer_name,
            'workspace_slug': workspace_slug,
            'sources_count': len(sources),
            'destinations_count': len(destinations_list),
            'audiences_count': len(all_audiences),
            'journeys_count': len([j for j in all_journeys if j.get('__typename') == 'Journey']),
            'campaigns_count': len([j for j in all_journeys if j.get('__typename') == 'Campaign']),
            'spaces_count': len(spaces)
        }

        with open(DATA_DIR / 'gateway_summary.json', 'w') as f:
            json.dump(summary, f, indent=2)

        audit_status['complete'] = True
        audit_status['progress'] = 100
        audit_status['message'] = 'Audit complete!'

    except Exception as e:
        audit_status['error'] = str(e)
        audit_status['message'] = f'Error: {str(e)}'
        print(f"Audit error: {e}")
    finally:
        audit_status['running'] = False

@app.route('/audit-status')
def get_audit_status():
    """Get current audit status"""
    return jsonify(audit_status)

@app.route('/progress')
def progress():
    """Progress page"""
    return render_template('gateway_progress.html')

@app.route('/dashboard')
def dashboard():
    """Main dashboard with both sources and audiences"""
    customer_name = session.get('customer_name', 'Customer')
    return render_template('gateway_dashboard.html', customer_name=customer_name, enable_experimental=ENABLE_EXPERIMENTAL_FEATURES)

@app.route('/sources')
def sources():
    """Sources view"""
    customer_name = session.get('customer_name', 'Customer')
    return render_template('gateway_sources.html', customer_name=customer_name, enable_experimental=ENABLE_EXPERIMENTAL_FEATURES)

@app.route('/audiences')
def audiences():
    """Audiences view"""
    customer_name = session.get('customer_name', 'Customer')
    return render_template('gateway_audiences.html', customer_name=customer_name, enable_experimental=ENABLE_EXPERIMENTAL_FEATURES)

@app.route('/destinations')
def destinations():
    """Destinations view"""
    customer_name = session.get('customer_name', 'Customer')
    return render_template('gateway_destinations.html', customer_name=customer_name, enable_experimental=ENABLE_EXPERIMENTAL_FEATURES)

@app.route('/audit-trail')
def audit_trail():
    """Audit trail view - governance and activity analysis"""
    customer_name = session.get('customer_name', 'Customer')
    return render_template('gateway_audit_trail.html', customer_name=customer_name, enable_experimental=ENABLE_EXPERIMENTAL_FEATURES)

@app.route('/api/destination-health-metrics', methods=['POST'])
def get_destination_health_metrics():
    """Fetch delivery health metrics for destinations on-demand"""
    try:
        gateway_token = session.get('gateway_token')
        workspace_slug = session.get('workspace_slug')

        if not gateway_token or not workspace_slug:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        destination_ids = data.get('destination_ids', [])
        days_back = data.get('days_back', 7)

        # Load destinations to get workspace ID
        destinations_file = DATA_DIR / 'gateway_destinations.json'
        if not destinations_file.exists():
            return jsonify({'error': 'No destinations data available'}), 404

        with open(destinations_file, 'r') as f:
            destinations_data = json.load(f)

        # Get workspace ID from summary or destinations data
        summary_file = DATA_DIR / 'gateway_summary.json'
        workspace_id = None
        if summary_file.exists():
            with open(summary_file, 'r') as f:
                summary = json.load(f)
                # Workspace ID might be in destinations data or we need to fetch it

        # Calculate time range
        from datetime import datetime, timedelta
        end_time = datetime.utcnow()
        start_time = end_time - timedelta(days=days_back)

        start_iso = start_time.isoformat() + 'Z'
        end_iso = end_time.isoformat() + 'Z'

        client = GatewayAPIClient(gateway_token, workspace_slug)

        # We need workspace ID - get it from workspace query
        workspace_data = client.get_workspace_connections()
        workspace_id = workspace_data.get('id')

        if not workspace_id:
            return jsonify({'error': 'Could not determine workspace ID'}), 400

        # Fetch metrics for each destination
        metrics_results = {}
        for dest_id in destination_ids:
            metrics = client.get_destination_delivery_metrics(
                dest_id,
                start_iso,
                end_iso,
                workspace_id
            )

            if metrics and metrics.get('deliveryOverview'):
                delivery = metrics['deliveryOverview']
                success_data = delivery.get('successfulDelivery', {})
                dataset = success_data.get('dataset', [])

                total_success = sum(d.get('total', 0) for d in dataset)
                total_retries = sum(d.get('totalRetryCount', 0) for d in dataset)

                # Find last delivery time
                last_delivery = None
                for ds in dataset:
                    series = ds.get('series', [])
                    if series:
                        last_time = max((s.get('time') for s in series if s.get('count', 0) > 0), default=None)
                        if last_time:
                            last_delivery = last_time

                metrics_results[dest_id] = {
                    'success_count': total_success,
                    'retry_count': total_retries,
                    'last_delivery': last_delivery,
                    'health_status': 'healthy' if total_retries < total_success * 0.05 else 'warning' if total_success > 0 else 'no_data'
                }
            else:
                metrics_results[dest_id] = {
                    'success_count': 0,
                    'retry_count': 0,
                    'last_delivery': None,
                    'health_status': 'no_data'
                }

        return jsonify({'metrics': metrics_results})

    except Exception as e:
        print(f"Error fetching health metrics: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/journeys')
def journeys():
    """Journeys view"""
    customer_name = session.get('customer_name', 'Customer')
    return render_template('gateway_journeys.html', customer_name=customer_name, enable_experimental=ENABLE_EXPERIMENTAL_FEATURES)

@app.route('/profile-insights')
def profile_insights():
    """Profile Insights view"""
    customer_name = session.get('customer_name', 'Customer')
    return render_template('gateway_profile_insights.html', customer_name=customer_name, enable_experimental=ENABLE_EXPERIMENTAL_FEATURES)

@app.route('/mtu')
def mtu():
    """MTU (Monthly Tracked Users) view"""
    customer_name = session.get('customer_name', 'Customer')
    return render_template('gateway_mtu.html', customer_name=customer_name, enable_experimental=ENABLE_EXPERIMENTAL_FEATURES)

@app.route('/api/audit-trail-data')
def get_audit_trail_data():
    """Get audit trail data for visualization"""
    try:
        audit_file = DATA_DIR / 'gateway_audit_trail.json'

        if not audit_file.exists():
            return jsonify({'error': 'No audit trail data available. Please run an audit first.'}), 404

        with open(audit_file, 'r') as f:
            audit_events = json.load(f)

        return jsonify({'events': audit_events})

    except Exception as e:
        print(f"Error loading audit trail: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/mtu-data')
def get_mtu_data():
    """Get MTU data for visualization"""
    try:
        mtu_file = DATA_DIR / 'gateway_mtu.json'

        if not mtu_file.exists():
            return jsonify({'error': 'No MTU data available. Please run an audit first.'}), 404

        with open(mtu_file, 'r') as f:
            mtu_data = json.load(f)

        return jsonify(mtu_data)

    except Exception as e:
        print(f"Error loading MTU data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-profile-insights-excel')
def export_profile_insights_excel():
    """Export profile insights data to Excel with multiple sheets"""
    try:
        workspace_slug = session.get('workspace_slug', 'workspace')

        # Load all profile insights CSVs
        identity_configs_file = DATA_DIR / 'gateway_profile_insights.csv'
        space_sources_file = DATA_DIR / 'gateway_space_sources.csv'

        if not identity_configs_file.exists():
            return jsonify({'error': 'No profile insights data available'}), 404

        # Read identity configs
        identity_configs = []
        with open(identity_configs_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            identity_configs = list(reader)

        # Read space sources
        space_sources = []
        if space_sources_file.exists():
            with open(space_sources_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                space_sources = list(reader)

        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Style headers
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Sheet 1: Identity Resolution Config
        ws_identity = wb.create_sheet('Identity Resolution')
        identity_headers = ['Workspace', 'Space', 'Space ID', 'ID Type', 'Priority', 'Limit', 'Seen']

        for col_idx, header in enumerate(identity_headers, 1):
            cell = ws_identity.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, config in enumerate(identity_configs, 2):
            ws_identity.cell(row=row_idx, column=1, value=config.get('Workspace', ''))
            ws_identity.cell(row=row_idx, column=2, value=config.get('Space', ''))
            ws_identity.cell(row=row_idx, column=3, value=config.get('Space ID', ''))
            ws_identity.cell(row=row_idx, column=4, value=config.get('ID Type', ''))
            ws_identity.cell(row=row_idx, column=5, value=config.get('Priority', ''))
            ws_identity.cell(row=row_idx, column=6, value=config.get('Limit', ''))
            ws_identity.cell(row=row_idx, column=7, value=config.get('Seen', ''))

        # Auto-size columns
        for col in ws_identity.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_identity.column_dimensions[column].width = min(max_length + 2, 50)

        # Sheet 2: Space Sources (Inbound)
        if space_sources:
            # Separate inbound and debugger sources
            inbound_sources = [s for s in space_sources if s.get('Type', '') != 'Personas']
            debugger_sources = [s for s in space_sources if s.get('Type', '') == 'Personas']

            # Inbound sources sheet
            if inbound_sources:
                ws_inbound = wb.create_sheet('Space Sources (Inbound)')
                inbound_headers = ['Workspace', 'Space', 'Space ID', 'Source Name', 'Status', 'Type', 'Category', 'Destinations']

                for col_idx, header in enumerate(inbound_headers, 1):
                    cell = ws_inbound.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for row_idx, source in enumerate(inbound_sources, 2):
                    ws_inbound.cell(row=row_idx, column=1, value=source.get('Workspace', ''))
                    ws_inbound.cell(row=row_idx, column=2, value=source.get('Space', ''))
                    ws_inbound.cell(row=row_idx, column=3, value=source.get('Space ID', ''))
                    ws_inbound.cell(row=row_idx, column=4, value=source.get('Source Name', ''))
                    ws_inbound.cell(row=row_idx, column=5, value=source.get('Status', ''))
                    ws_inbound.cell(row=row_idx, column=6, value=source.get('Type', ''))
                    ws_inbound.cell(row=row_idx, column=7, value=source.get('Category', ''))
                    ws_inbound.cell(row=row_idx, column=8, value=source.get('Destinations', ''))

                # Auto-size columns
                for col in ws_inbound.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws_inbound.column_dimensions[column].width = min(max_length + 2, 50)

            # Debugger sources sheet
            if debugger_sources:
                ws_debugger = wb.create_sheet('Profile Debugger Sources')
                debugger_headers = ['Workspace', 'Space', 'Space ID', 'Source Name', 'Status', 'Destinations']

                for col_idx, header in enumerate(debugger_headers, 1):
                    cell = ws_debugger.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for row_idx, source in enumerate(debugger_sources, 2):
                    ws_debugger.cell(row=row_idx, column=1, value=source.get('Workspace', ''))
                    ws_debugger.cell(row=row_idx, column=2, value=source.get('Space', ''))
                    ws_debugger.cell(row=row_idx, column=3, value=source.get('Space ID', ''))
                    ws_debugger.cell(row=row_idx, column=4, value=source.get('Source Name', ''))
                    ws_debugger.cell(row=row_idx, column=5, value=source.get('Status', ''))
                    ws_debugger.cell(row=row_idx, column=6, value=source.get('Destinations', ''))

                # Auto-size columns
                for col in ws_debugger.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws_debugger.column_dimensions[column].width = min(max_length + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f"{workspace_slug}_profile_insights_{date_str}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error generating profile insights Excel: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/audit_data/<path:filename>')
def serve_data(filename):
    """Serve audit data files"""
    file_path = DATA_DIR / filename
    if file_path.exists():
        from flask import send_file
        return send_file(file_path)
    return f"File not found: {filename}", 404

@app.route('/api/source-schema/<source_slug>')
def get_source_schema_api(source_slug):
    """Get source event schema via Gateway API"""
    try:
        gateway_token = session.get('gateway_token')
        workspace_slug = session.get('workspace_slug')

        if not gateway_token or not workspace_slug:
            return jsonify({'error': 'Session expired. Please run a new audit.'}), 401

        client = GatewayAPIClient(gateway_token, workspace_slug)
        schema_data = client.get_source_schema(source_slug)

        return jsonify(schema_data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/audience-definition/<space_id>/<audience_id>')
def get_audience_definition_api(space_id, audience_id):
    """Get audience definition via Gateway API"""
    try:
        gateway_token = session.get('gateway_token')
        workspace_slug = session.get('workspace_slug')

        if not gateway_token or not workspace_slug:
            return jsonify({'error': 'Session expired. Please run a new audit.'}), 401

        client = GatewayAPIClient(gateway_token, workspace_slug)
        definition_data = client.get_audience_definition(space_id, audience_id)

        return jsonify(definition_data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-sources-excel')
def export_sources_excel():
    """Export sources with event details to Excel - LLM-optimized format"""
    try:
        gateway_token = session.get('gateway_token')
        workspace_slug = session.get('workspace_slug')

        if not gateway_token or not workspace_slug:
            return jsonify({'error': 'Session expired. Please run a new audit.'}), 401

        # Load sources data
        sources_file = DATA_DIR / 'gateway_sources.csv'
        if not sources_file.exists():
            return jsonify({'error': 'No source data available'}), 404

        sources = []
        with open(sources_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            sources = list(reader)

        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Helper function to extract environment from labels
        def extract_environment(labels_str):
            """Extract environment value from labels like 'environment=media-dev'"""
            if not labels_str:
                return ''
            for label in labels_str.split(','):
                if 'environment=' in label.lower():
                    return label.split('=')[1].strip()
            return ''

        # Helper function to determine source health
        def get_source_health(status, total_allowed):
            """Derive health status from source state and volume"""
            if status == 'DISABLED':
                return 'DISABLED'
            elif total_allowed > 0:
                return 'ACTIVE'
            else:
                return 'NO_RECENT_DATA'

        # Create main Sources Summary sheet
        ws_main = wb.create_sheet('Sources Summary')

        # Main sheet headers - LLM-friendly normalized columns
        main_headers = ['workspace', 'source_name', 'source_slug', 'source_status', 'source_health',
                       'environment', 'source_type', 'technical_type',
                       'connected_destinations', 'connected_warehouses', 'labels',
                       'total_events', 'traits_count', 'total_allowed_7d', 'total_blocked_7d',
                       'has_recent_data', 'volume_window']

        # Style headers
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, header in enumerate(main_headers, 1):
            cell = ws_main.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Client for fetching schemas
        client = GatewayAPIClient(gateway_token, workspace_slug)

        # Collect all events and traits for master sheet
        all_master_rows = []

        # Process each source
        for row_idx, source in enumerate(sources, 2):
            source_name = source.get('Name', '')
            source_slug = source.get('Slug', '')
            source_status = source.get('Status', '')
            labels = source.get('Labels', '')
            environment = extract_environment(labels)

            # Track for main summary
            traits_list = [t.strip() for t in source.get('Identify Traits', '').split(',') if t.strip()]

            # Fetch schema data for this source
            if source_slug:
                try:
                    schema_data = client.get_source_schema(source_slug)

                    # Build event array (same logic as frontend)
                    recent_events_map = {}
                    for e in schema_data.get('events', []):
                        recent_events_map[e['name']] = {
                            'type': e['type'],
                            'allowed': e.get('counts', {}).get('allowed', 0),
                            'denied': e.get('counts', {}).get('denied', 0)
                        }

                    all_events_array = []

                    # Collect events from collections
                    for collection in schema_data.get('collections', []):
                        for event in collection.get('events', []):
                            recent_data = recent_events_map.get(event['name'])
                            all_events_array.append({
                                'name': event['name'],
                                'type': 'TRACK',
                                'isPlanned': event.get('isPlanned', False),
                                'isRecent': recent_data is not None,
                                'allowed': recent_data['allowed'] if recent_data else 0,
                                'denied': recent_data['denied'] if recent_data else 0
                            })

                    # Add recent events not in schema
                    for recent_event in schema_data.get('events', []):
                        if not any(e['name'] == recent_event['name'] for e in all_events_array):
                            all_events_array.append({
                                'name': recent_event['name'],
                                'type': recent_event['type'],
                                'isPlanned': False,
                                'isRecent': True,
                                'allowed': recent_event.get('counts', {}).get('allowed', 0),
                                'denied': recent_event.get('counts', {}).get('denied', 0)
                            })

                    total_events = len(all_events_array)
                    total_allowed = sum(e['allowed'] for e in all_events_array)
                    total_blocked = sum(e['denied'] for e in all_events_array)

                    ws_main.cell(row=row_idx, column=10, value=total_events)
                    ws_main.cell(row=row_idx, column=11, value=len(traits_list))
                    ws_main.cell(row=row_idx, column=12, value=total_allowed)
                    ws_main.cell(row=row_idx, column=13, value=total_blocked)

                    # Only create detail sheets if there are events with volume
                    events_with_volume = [e for e in all_events_array if e['allowed'] > 0 or e['denied'] > 0]

                    if events_with_volume:
                        # Create Events sheet for this source
                        safe_sheet_name = source_name[:28] + '_E' if len(source_name) > 28 else source_name + '_Events'
                        safe_sheet_name = ''.join(c for c in safe_sheet_name if c.isalnum() or c in (' ', '_', '-'))[:31]

                        ws_events = wb.create_sheet(safe_sheet_name)

                        # Events sheet headers
                        event_headers = ['Event Name', 'Type', 'Status', 'Allowed (7d)', 'Blocked (7d)']
                        for col_idx, header in enumerate(event_headers, 1):
                            cell = ws_events.cell(row=1, column=col_idx, value=header)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')

                        # Sort events by volume (allowed desc)
                        events_with_volume.sort(key=lambda x: x['allowed'], reverse=True)

                        for evt_row, event in enumerate(events_with_volume, 2):
                            event_type = event['type']
                            if event_type == 'TRACK' and event['name'] == 'Page Viewed':
                                event_type = 'PAGE'

                            status = 'Planned' if event['isPlanned'] else 'Unplanned'

                            ws_events.cell(row=evt_row, column=1, value=event['name'])
                            ws_events.cell(row=evt_row, column=2, value=event_type)
                            ws_events.cell(row=evt_row, column=3, value=status)
                            ws_events.cell(row=evt_row, column=4, value=event['allowed'])
                            ws_events.cell(row=evt_row, column=5, value=event['denied'])

                        # Auto-size columns
                        for col in ws_events.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            ws_events.column_dimensions[column].width = min(max_length + 2, 50)

                    # Create Traits sheet if traits exist - with volume data
                    if traits_list:
                        # Fetch trait volume from objectProperties in "users" collection
                        traits_with_volume = []
                        for collection in schema_data.get('collections', []):
                            if collection.get('name', '').lower() == 'users':
                                for prop in collection.get('objectProperties', []):
                                    if prop.get('enabled', True) and not prop.get('archived', False):
                                        trait_key = prop.get('key', '')
                                        stats = prop.get('stats', {})
                                        allowed = stats.get('allowed', 0)
                                        denied = stats.get('denied', 0)
                                        # Only include traits with volume
                                        if allowed > 0 or denied > 0:
                                            traits_with_volume.append({
                                                'key': trait_key,
                                                'allowed': allowed,
                                                'denied': denied
                                            })
                                break

                        # Only create sheet if we have traits with actual volume
                        if traits_with_volume:
                            safe_sheet_name = source_name[:28] + '_T' if len(source_name) > 28 else source_name + '_Traits'
                            safe_sheet_name = ''.join(c for c in safe_sheet_name if c.isalnum() or c in (' ', '_', '-'))[:31]

                            ws_traits = wb.create_sheet(safe_sheet_name)

                            trait_headers = ['Trait Name', 'Allowed (7d)', 'Blocked (7d)']
                            for col_idx, header in enumerate(trait_headers, 1):
                                cell = ws_traits.cell(row=1, column=col_idx, value=header)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center')

                            # Sort traits by volume (allowed desc)
                            traits_with_volume.sort(key=lambda x: x['allowed'], reverse=True)

                            for trait_row, trait in enumerate(traits_with_volume, 2):
                                ws_traits.cell(row=trait_row, column=1, value=trait['key'])
                                ws_traits.cell(row=trait_row, column=2, value=trait['allowed'])
                                ws_traits.cell(row=trait_row, column=3, value=trait['denied'])

                            # Auto-size columns
                            for col in ws_traits.columns:
                                max_length = 0
                                column = col[0].column_letter
                                for cell in col:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                ws_traits.column_dimensions[column].width = min(max_length + 2, 50)

                except Exception as e:
                    print(f"Error fetching schema for {source_slug}: {e}")
                    ws_main.cell(row=row_idx, column=10, value=0)
                    ws_main.cell(row=row_idx, column=11, value=len(traits_list))
                    ws_main.cell(row=row_idx, column=12, value=0)
                    ws_main.cell(row=row_idx, column=13, value=0)
            else:
                ws_main.cell(row=row_idx, column=10, value=0)
                ws_main.cell(row=row_idx, column=11, value=len(traits_list))
                ws_main.cell(row=row_idx, column=12, value=0)
                ws_main.cell(row=row_idx, column=13, value=0)

        # Auto-size main sheet columns
        for col in ws_main.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_main.column_dimensions[column].width = min(max_length + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f"{workspace_slug}_sources_detailed_{date_str}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error generating Excel: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-sources-excel-v2')
def export_sources_excel_v2():
    """Export sources - LLM-optimized with flattened master sheet"""
    try:
        gateway_token = session.get('gateway_token')
        workspace_slug = session.get('workspace_slug')

        if not gateway_token or not workspace_slug:
            return jsonify({'error': 'Session expired'}), 401

        # Load sources
        sources_file = DATA_DIR / 'gateway_sources.csv'
        if not sources_file.exists():
            return jsonify({'error': 'No source data'}), 404

        sources = []
        with open(sources_file, 'r', encoding='utf-8') as f:
            sources = list(csv.DictReader(f))

        # Helper functions
        def extract_environment(labels_str):
            if not labels_str:
                return ''
            for label in labels_str.split(','):
                if 'environment=' in label.lower():
                    return label.split('=')[1].strip()
            return ''

        def get_source_health(status, total_allowed):
            if status == 'DISABLED':
                return 'DISABLED'
            return 'ACTIVE' if total_allowed > 0 else 'NO_RECENT_DATA'

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Style
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        client = GatewayAPIClient(gateway_token, workspace_slug)

        # Master flattened table - all events and traits in one sheet
        ws_master = wb.create_sheet('Master Data')
        master_headers = ['workspace', 'source_name', 'source_slug', 'source_status', 'source_health',
                         'environment', 'source_type', 'technical_type', 'record_type', 'object_name',
                         'object_type', 'planning_status', 'allowed_7d', 'blocked_7d',
                         'volume_window', 'has_recent_data']

        for col_idx, header in enumerate(master_headers, 1):
            cell = ws_master.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        master_row_idx = 2

        # Sources Summary sheet
        ws_summary = wb.create_sheet('Sources Summary')
        summary_headers = ['workspace', 'source_name', 'source_slug', 'source_status', 'source_health',
                          'environment', 'source_type', 'technical_type',
                          'connected_destinations', 'connected_warehouses',
                          'total_events', 'traits_count', 'total_allowed_7d', 'total_blocked_7d',
                          'has_recent_data', 'volume_window']

        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Process each source
        for source in sources:
            source_name = source.get('Name', '')
            source_slug = source.get('Slug', '')
            source_status = source.get('Status', '')
            source_type = source.get('Type', '')
            technical_type = source.get('Technical Type', '')
            labels = source.get('Labels', '')
            environment = extract_environment(labels)

            total_allowed = 0
            total_blocked = 0
            total_events = 0
            traits_count = 0

            # Fetch schema
            if source_slug:
                try:
                    schema_data = client.get_source_schema(source_slug)

                    # Process events
                    recent_events_map = {}
                    for e in schema_data.get('events', []):
                        recent_events_map[e['name']] = {
                            'type': e['type'],
                            'allowed': e.get('counts', {}).get('allowed', 0),
                            'denied': e.get('counts', {}).get('denied', 0)
                        }

                    all_events = []
                    for collection in schema_data.get('collections', []):
                        for event in collection.get('events', []):
                            recent = recent_events_map.get(event['name'])
                            all_events.append({
                                'name': event['name'],
                                'type': 'TRACK',
                                'isPlanned': event.get('isPlanned', False),
                                'allowed': recent['allowed'] if recent else 0,
                                'denied': recent['denied'] if recent else 0
                            })

                    # Add unplanned recent events
                    for name, data in recent_events_map.items():
                        if not any(e['name'] == name for e in all_events):
                            all_events.append({
                                'name': name,
                                'type': data['type'],
                                'isPlanned': False,
                                'allowed': data['allowed'],
                                'denied': data['denied']
                            })

                    total_events = len(all_events)
                    total_allowed = sum(e['allowed'] for e in all_events)
                    total_blocked = sum(e['denied'] for e in all_events)

                    # Add events to master sheet (keep zeros)
                    for event in all_events:
                        event_type = event['type']
                        if event_type == 'TRACK' and event['name'] == 'Page Viewed':
                            event_type = 'PAGE'

                        planning_status = 'Planned' if event['isPlanned'] else 'Unplanned'
                        allowed = event['allowed']
                        blocked = event['denied']

                        ws_master.cell(row=master_row_idx, column=1, value=workspace_slug)
                        ws_master.cell(row=master_row_idx, column=2, value=source_name)
                        ws_master.cell(row=master_row_idx, column=3, value=source_slug)
                        ws_master.cell(row=master_row_idx, column=4, value=source_status)
                        ws_master.cell(row=master_row_idx, column=5, value=get_source_health(source_status, total_allowed))
                        ws_master.cell(row=master_row_idx, column=6, value=environment)
                        ws_master.cell(row=master_row_idx, column=7, value=source_type)
                        ws_master.cell(row=master_row_idx, column=8, value=technical_type)
                        ws_master.cell(row=master_row_idx, column=9, value='event')
                        ws_master.cell(row=master_row_idx, column=10, value=event['name'])
                        ws_master.cell(row=master_row_idx, column=11, value=event_type)
                        ws_master.cell(row=master_row_idx, column=12, value=planning_status)
                        ws_master.cell(row=master_row_idx, column=13, value=allowed)
                        ws_master.cell(row=master_row_idx, column=14, value=blocked)
                        ws_master.cell(row=master_row_idx, column=15, value='last_7_days')
                        ws_master.cell(row=master_row_idx, column=16, value='TRUE' if allowed > 0 else 'FALSE')
                        master_row_idx += 1

                    # Process traits
                    for collection in schema_data.get('collections', []):
                        if collection.get('name', '').lower() == 'users':
                            for prop in collection.get('objectProperties', []):
                                if prop.get('enabled', True) and not prop.get('archived', False):
                                    trait_key = prop.get('key', '')
                                    stats = prop.get('stats', {})
                                    allowed = stats.get('allowed', 0)
                                    blocked = stats.get('denied', 0)
                                    traits_count += 1

                                    # Add trait to master (keep zeros)
                                    ws_master.cell(row=master_row_idx, column=1, value=workspace_slug)
                                    ws_master.cell(row=master_row_idx, column=2, value=source_name)
                                    ws_master.cell(row=master_row_idx, column=3, value=source_slug)
                                    ws_master.cell(row=master_row_idx, column=4, value=source_status)
                                    ws_master.cell(row=master_row_idx, column=5, value=get_source_health(source_status, total_allowed))
                                    ws_master.cell(row=master_row_idx, column=6, value=environment)
                                    ws_master.cell(row=master_row_idx, column=7, value=source_type)
                                    ws_master.cell(row=master_row_idx, column=8, value=technical_type)
                                    ws_master.cell(row=master_row_idx, column=9, value='trait')
                                    ws_master.cell(row=master_row_idx, column=10, value=trait_key)
                                    ws_master.cell(row=master_row_idx, column=11, value='')
                                    ws_master.cell(row=master_row_idx, column=12, value='')
                                    ws_master.cell(row=master_row_idx, column=13, value=allowed)
                                    ws_master.cell(row=master_row_idx, column=14, value=blocked)
                                    ws_master.cell(row=master_row_idx, column=15, value='last_7_days')
                                    ws_master.cell(row=master_row_idx, column=16, value='TRUE' if allowed > 0 else 'FALSE')
                                    master_row_idx += 1
                            break

                except Exception as e:
                    print(f"Error fetching schema for {source_slug}: {e}")

            # Write to summary sheet
            summary_row = [
                workspace_slug, source_name, source_slug, source_status,
                get_source_health(source_status, total_allowed), environment,
                source_type, technical_type,
                source.get('Connected Destinations', ''),
                source.get('Connected Warehouses', ''),
                total_events, traits_count, total_allowed, total_blocked,
                'TRUE' if total_allowed > 0 else 'FALSE', 'last_7_days'
            ]

            summary_row_idx = sources.index(source) + 2  # +2 for header row and 1-indexed
            for col_idx, value in enumerate(summary_row, 1):
                ws_summary.cell(row=summary_row_idx, column=col_idx, value=value)

        # Auto-size all columns
        for ws in [ws_master, ws_summary]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{workspace_slug}_sources_llm_optimized_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name=filename)

    except Exception as e:
        print(f"Error generating LLM-optimized Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/recommendations')
def recommendations_page():
    """Recommendations page (experimental feature)"""
    if not ENABLE_EXPERIMENTAL_FEATURES:
        return "Feature not available", 404

    workspace_slug = session.get('workspace_slug')
    if not workspace_slug:
        return redirect('/')
    return render_template('recommendations.html', workspace_slug=workspace_slug)

@app.route('/api/generate-recommendations', methods=['POST'])
def generate_recommendations_api():
    """Generate goal-driven workspace recommendations using targeted AI analysis (experimental)"""
    if not ENABLE_EXPERIMENTAL_FEATURES:
        return jsonify({'success': False, 'error': 'Feature not available'}), 404

    try:
        from recommendations_cache import RecommendationsCache
        from data_structurer import DataStructurer
        from business_inference_prompts import BusinessInferencePrompts
        from goal_driven_prompts import GoalDrivenPrompts
        from enhanced_audit_prompts import EnhancedAuditPrompts  # New enhanced system
        # from mcp_collective_intelligence import MCPCollectiveIntelligence  # Disabled - no database
        from gemini_client import GeminiClient
        import re

        # Initialize cache
        cache = RecommendationsCache(str(DATA_DIR))

        # Get workspace slug from session
        workspace_slug = session.get('workspace_slug', 'unknown')

        # Get request parameters (goal-driven approach)
        req_data = request.json or {}
        goal = req_data.get('goal')
        output_type = req_data.get('output_type', 'recommended_actions')  # Default to full JSON
        industry_override = req_data.get('industry_override')
        business_model_override = req_data.get('business_model_override')
        user_notes = req_data.get('user_notes', '')
        force_refresh = req_data.get('force_refresh', False)

        if not goal:
            return jsonify({
                'success': False,
                'error': 'Goal is required'
            }), 400

        # Simplified: output_type no longer required (always return full JSON)
        import hashlib
        context_hash = hashlib.md5(f"{industry_override}_{business_model_override}_{user_notes}".encode()).hexdigest()[:8]
        cache_key = f"{goal}_{context_hash}"

        # Only use cache if not forcing refresh
        if not force_refresh:
            cached = cache.get_cached_recommendations(workspace_slug, cache_key)
            if cached:
                print("✓ Using cached result")
                return jsonify({
                    'success': True,
                    'result': cached.get('result'),
                    'layer0': cached.get('layer0'),
                    'cached': True
                })
        else:
            print("🔄 Force refresh - bypassing cache")

        # Check rate limits before calling Gemini
        allowed, reason = cache.check_rate_limit()
        if not allowed:
            return jsonify({
                'success': False,
                'error': reason,
                'rate_limited': True
            }), 429

        # Record API call
        cache.record_api_call()

        print(f"🎯 Goal-driven analysis: {goal} -> {output_type}")

        # Step 1: Generate findings (needed for data structuring)
        print("🔍 Generating rule-based findings...")
        from recommendations_engine import generate_recommendations
        findings = generate_recommendations(str(DATA_DIR))

        # Step 2: Structure data for AI
        print("📊 Structuring workspace data...")
        structurer = DataStructurer(str(DATA_DIR))
        structured_data = structurer.structure_for_gemini(findings)

        # Step 3: Business context (use overrides or simple default - SKIP AI inference to save API calls)
        print("🔍 Setting business context...")

        # Get Gemini API key
        gemini_api_key = os.environ.get('GEMINI_API_KEY')
        if not gemini_api_key:
            return jsonify({
                'success': False,
                'error': 'GEMINI_API_KEY environment variable not set'
            }), 500

        # Initialize Gemini client (handles SSL bypass in dev mode)
        gemini_client = GeminiClient(gemini_api_key)

        if industry_override and business_model_override:
            print(f"   Using overrides: {industry_override} / {business_model_override}")
            layer0_result = {
                'industry': {'primary': industry_override, 'primary_confidence': 'high'},
                'business_model': {'primary': business_model_override, 'primary_confidence': 'high'}
            }
        else:
            # Use simple default to avoid extra API call
            # User can override in the UI if needed
            workspace_slug = session.get('workspace_slug', 'unknown')
            print(f"   Using default context for {workspace_slug}")
            layer0_result = {
                'industry': {'primary': 'Media/Publishing', 'primary_confidence': 'medium'},
                'business_model': {'primary': 'Subscription', 'primary_confidence': 'medium'}
            }

        # Step 4: Query MCP for collective intelligence (DISABLED FOR NOW)
        print("🌐 Skipping collective intelligence (database disabled)...")
        # mcp = MCPCollectiveIntelligence()
        # collective_insights = mcp.get_contextual_insights(layer0_result)
        collective_insights = {
            'industry': layer0_result.get('industry', {}).get('primary', 'Unknown'),
            'business_model': layer0_result.get('business_model', {}).get('primary', 'Unknown'),
            'similar_workspaces_analyzed': 0,
            'benchmarks': {},
            'best_practices': [],
            'collective_context': ''
        }

        # Step 5: Build business context string
        business_context = f"""
Industry: {layer0_result.get('industry', {}).get('primary', 'Unknown')}
Business Model: {layer0_result.get('business_model', {}).get('primary', 'Unknown')}

{collective_insights.get('collective_context', '')}
"""

        # Step 6: Select goal-specific prompt
        print(f"📝 Generating {goal} prompt for {output_type}...")

        # Use enhanced prompts for workspace_audit, fallback to original for others
        if goal == 'workspace_audit':
            print(f"   ✨ Using ENHANCED prompt system with confidence levels ✨")
            enhanced_prompts = EnhancedAuditPrompts()
            prompt = enhanced_prompts.workspace_audit_with_confidence_levels(
                structured_data,
                business_context,
                user_notes
            )
        else:
            # Fallback to original system for growth_usecases and activation_expansion
            print(f"   Using original prompt system")
            prompts = GoalDrivenPrompts()
            if goal == 'growth_usecases':
                prompt = prompts.goal_growth_usecases(structured_data, business_context, user_notes, 'recommended_actions')
            elif goal == 'activation_expansion':
                prompt = prompts.goal_activation_expansion(structured_data, business_context, user_notes, 'recommended_actions')
            else:
                return jsonify({
                    'success': False,
                    'error': f'Unknown goal: {goal}'
                }), 400

        # Step 7: Call Gemini with goal-specific prompt
        print(f"✨ Calling Gemini for {goal} analysis...")
        print(f"   Prompt length: {len(prompt):,} characters")

        # Use staged summarization if prompt is too large (>30k chars)
        USE_STAGED_SUMMARIZATION = len(prompt) > 30000

        if USE_STAGED_SUMMARIZATION:
            print(f"   ⚠️  Large prompt detected ({len(prompt):,} chars)")
            print(f"   🔄 Using staged summarization to reduce API load...")
            try:
                # This will break the analysis into 4 smaller API calls
                staged_result = gemini_client.staged_summarization(
                    structured_data, goal, business_context, model='gemini-2.5-flash'
                )
                # For now, we still need to do final generation with compact data
                # TODO: Update prompts to use compact summaries
                print(f"   ⚠️  Staged summarization complete, but falling back to full prompt for now")
            except Exception as stage_error:
                print(f"   ⚠️  Staged summarization failed: {stage_error}")
                print(f"   🔄 Falling back to single large prompt...")

        # Using gemini-2.5-flash with intelligent retry and fallback
        response_text = gemini_client.generate_content(prompt, model='gemini-2.5-flash').strip()
        print(f"   Response length: {len(response_text)} characters")
        print(f"   Response preview: {response_text[:200]}")

        # Always expect JSON structure (simplified)
        # Try direct JSON parse first
        try:
            result = json.loads(response_text)
            print(f"   ✓ Direct JSON parse successful, keys: {list(result.keys())}")
        except json.JSONDecodeError as e:
            print(f"   ⚠️ Direct JSON parse failed: {e}")
            # Try extracting from markdown code block
            # Look for ```json or ``` followed by JSON content
            if response_text.strip().startswith('```'):
                # Remove opening ``` and optional json marker
                json_str = response_text.strip()
                json_str = re.sub(r'^```(?:json)?\s*', '', json_str)
                # Remove closing ```
                json_str = re.sub(r'\s*```\s*$', '', json_str)

                print(f"   Extracted from markdown block: {len(json_str)} chars")
                try:
                    result = json.loads(json_str)
                    print(f"   ✓ Markdown JSON parse successful, keys: {list(result.keys())}")
                except json.JSONDecodeError as e2:
                    print(f"   ✗ Markdown JSON parse failed: {e2}")
                    print(f"   First 200 chars of extracted JSON: {json_str[:200]}")
                    return jsonify({
                        'success': False,
                        'error': f'Could not parse JSON from markdown: {str(e2)}',
                        'response_preview': response_text[:500]
                    }), 500
            else:
                # If still no JSON, return error with response preview
                print(f"⚠️ No JSON found in response. First 500 chars: {response_text[:500]}")
                return jsonify({
                    'success': False,
                    'error': 'Gemini returned non-JSON response',
                    'response_preview': response_text[:500]
                }), 500

        # Step 8: Contribute to MCP for future learning (DISABLED)
        # Database functionality disabled per user request
        print("📝 Skipping MCP contribution (database disabled)")

        # Prepare response
        response_data = {
            'success': True,
            'result': result,
            'layer0': layer0_result,
            'collective_insights': collective_insights,
            'cached': False
        }

        # Cache the result
        cache.cache_recommendations(workspace_slug, response_data, cache_key)

        return jsonify(response_data)

    except Exception as e:
        print(f"❌ Error generating goal-driven recommendations: {e}")
        import traceback
        traceback.print_exc()

        # Classify error type for better user messaging
        error_msg = str(e)

        # Rate limit errors
        if 'Rate limited' in error_msg or '429' in error_msg:
            return jsonify({
                'success': False,
                'rate_limited': True,
                'error': 'Gemini API rate limit reached. The system will automatically retry with exponential backoff. If this persists, please wait 60 seconds before trying again.',
                'user_message': '⏱️ Rate limit reached. Retrying automatically...'
            }), 429

        # Service unavailable errors
        elif '503' in error_msg or '504' in error_msg or 'unavailable' in error_msg.lower() or 'timeout' in error_msg.lower():
            return jsonify({
                'success': False,
                'service_unavailable': True,
                'error': error_msg,
                'user_message': '🔧 Gemini API is temporarily unavailable. The system attempted multiple retries and fallback models. Please try again in a few minutes.'
            }), 503

        # All other errors
        return jsonify({
            'success': False,
            'error': error_msg,
            'user_message': f'❌ An error occurred: {error_msg[:200]}'
        }), 500

@app.route('/api/recommendations-status')
def recommendations_status():
    """Get current rate limit status"""
    try:
        from recommendations_cache import RecommendationsCache
        cache = RecommendationsCache(str(DATA_DIR))
        status = cache.get_rate_limit_status()
        return jsonify(status)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-recommendations-csv')
def export_recommendations_csv():
    """Export recommendations as CSV"""
    try:
        from recommendations_engine import generate_recommendations
        from export_manager import ExportManager

        # Generate recommendations
        findings = generate_recommendations(str(DATA_DIR))

        # Export to CSV
        exporter = ExportManager(str(DATA_DIR))
        csv_data = exporter.export_recommendations_csv(findings)

        # Return as downloadable file
        from flask import Response
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=recommendations.csv'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-sources-destinations-csv')
def export_sources_destinations_csv():
    """Export sources with destinations as CSV"""
    try:
        from export_manager import ExportManager

        exporter = ExportManager(str(DATA_DIR))
        csv_data = exporter.export_sources_with_destinations_csv()

        from flask import Response
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=sources_with_destinations.csv'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-audiences-destinations-csv')
def export_audiences_destinations_csv():
    """Export audiences with destinations as CSV"""
    try:
        from export_manager import ExportManager

        exporter = ExportManager(str(DATA_DIR))
        csv_data = exporter.export_audiences_with_destinations_csv()

        from flask import Response
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=audiences_with_destinations.csv'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-all-audit-data')
def export_all_audit_data():
    """Export all audit data as ZIP"""
    try:
        from recommendations_engine import generate_recommendations
        from export_manager import ExportManager

        # Generate recommendations
        findings = generate_recommendations(str(DATA_DIR))

        # Export all as ZIP
        exporter = ExportManager(str(DATA_DIR))
        zip_data = exporter.export_all_as_zip(findings)

        from flask import Response
        workspace_slug = session.get('workspace_slug', 'workspace')
        filename = f'segment_audit_{workspace_slug}_{datetime.now().strftime("%Y%m%d")}.zip'

        return Response(
            zip_data,
            mimetype='application/zip',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/reset')
def reset():
    """Reset and start new audit"""
    session.clear()
    return redirect('/')

if __name__ == '__main__':
    print("="*80)
    print("Gateway API Audit Dashboard (Testing Only)")
    print("="*80)
    print()
    print("🚀 Starting server...")
    print(f"📁 Data directory: {DATA_DIR}")
    print()
    print("📊 Dashboard URL: http://localhost:5003")
    print()
    print("💡 This version uses ONLY Gateway API (no Public API)")
    print("💡 Press Ctrl+C to stop the server")
    print("="*80)
    print()

    app.run(host='0.0.0.0', port=5003, debug=False)
